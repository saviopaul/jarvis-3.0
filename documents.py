"""
documents.py - Universal Document Parser & Scanned File OCR Engine

Parses, OCRs, and extracts text/data from:
- Scanned & Digital PDFs (.pdf) via native Gemini Vision OCR + PyPDF
- Scanned Images & Photos (.jpg, .jpeg, .png, .webp, .tiff)
- Word Documents (.docx)
- Excel Spreadsheets (.xlsx, .xls)
- CSV / TSV (.csv, .tsv)
- Code & Text (.txt, .md, .py, .js, .json, .html, .sql, .yaml)
"""

import os
import io
import csv
import json
import base64
import logging
import requests

logger = logging.getLogger(__name__)


def parse_document(file_bytes: bytes, file_name: str) -> str:
    """Extracts text content from various document formats (digital or scanned)."""
    ext = os.path.splitext(file_name.lower())[1]
    
    try:
        # 1. Plain Text / Code / Markdown / JSON / Configs
        if ext in [".txt", ".md", ".py", ".js", ".html", ".css", ".json", ".yaml", ".yml", ".log", ".sql", ".sh", ".env", ".xml"]:
            return file_bytes.decode("utf-8", errors="ignore")

        # 2. CSV / TSV Files
        elif ext in [".csv", ".tsv"]:
            delimiter = "\t" if ext == ".tsv" else ","
            content = file_bytes.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(content), delimiter=delimiter)
            rows = [", ".join(row) for row in list(reader)[:150]]  # Top 150 rows
            return f"CSV/TSV Data Preview ({file_name}):\n" + "\n".join(rows)

        # 3. PDF Files (Digital + Scanned OCR)
        elif ext == ".pdf":
            try:
                import pypdf
                reader = pypdf.PdfReader(io.BytesIO(file_bytes))
                text = ""
                for idx, page in enumerate(reader.pages[:30]):  # Up to 30 pages
                    extracted = page.extract_text()
                    if extracted:
                        text += f"\n--- Page {idx+1} ---\n" + extracted
                
                # If digital text extraction was successful and has substance
                if len(text.strip()) > 60:
                    return text
            except Exception as e:
                logger.warning(f"PyPDF extraction error: {e}, falling back to Gemini Vision OCR")

            # Fallback: Scanned PDF -> Perform Gemini Native Multimodal PDF OCR
            return _ocr_pdf_with_gemini(file_bytes, file_name)

        # 4. Word Documents (.docx)
        elif ext == ".docx":
            try:
                import docx
                doc = docx.Document(io.BytesIO(file_bytes))
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                return f"Word Document ({file_name}):\n\n" + "\n".join(paragraphs)
            except Exception as e:
                logger.error(f"DOCX extraction error: {e}")
                return f"Error reading Word document: {str(e)}"

        # 5. Excel Spreadsheets (.xlsx, .xls)
        elif ext in [".xlsx", ".xls"]:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                output = f"Excel Spreadsheet ({file_name}):\n"
                for sheet_name in wb.sheetnames[:5]:
                    sheet = wb[sheet_name]
                    output += f"\n📊 Sheet: {sheet_name}\n"
                    for row in sheet.iter_rows(values_only=True, max_row=100):
                        non_empty = [str(cell) for cell in row if cell is not None]
                        if non_empty:
                            output += " | ".join(non_empty) + "\n"
                return output
            except Exception as e:
                logger.error(f"Excel extraction error: {e}")
                return f"Error reading Excel file: {str(e)}"

        # 6. Fallback: Try decoding as raw text
        else:
            return file_bytes.decode("utf-8", errors="ignore")

    except Exception as e:
        return f"Could not extract text from {file_name}: {str(e)}"


def _ocr_pdf_with_gemini(file_bytes: bytes, file_name: str) -> str:
    """Uses Gemini 2.5 Flash native PDF vision OCR to extract text from scanned PDFs."""
    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        return "Scanned PDF detected, but GEMINI_API_KEY is not configured for OCR."
        
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
    b64_data = base64.b64encode(file_bytes).decode("utf-8")
    
    prompt = (
        f"You are the Lead High-Accuracy OCR & Document Transcription Engine.\n"
        f"This PDF '{file_name}' contains scanned or image-based pages.\n"
        f"1. Transcribe ALL visible text, tables, figures, numbers, and handwritten notes verbatim.\n"
        f"2. Maintain table layouts and column alignments.\n"
        f"3. Provide a clean, copyable transcription block."
    )
    
    payload = {
        "contents": [
            {
                "parts": [
                    {"text": prompt},
                    {
                        "inline_data": {
                            "mime_type": "application/pdf",
                            "data": b64_data
                        }
                    }
                ]
            }
        ]
    }
    
    try:
        resp = requests.post(url, json=payload, timeout=50)
        data = resp.json()
        if "candidates" in data and len(data["candidates"]) > 0:
            return data["candidates"][0]["content"]["parts"][0]["text"]
        return f"OCR processing failed: {data.get('error', {}).get('message', resp.text)}"
    except Exception as e:
        return f"Scanned PDF OCR error: {str(e)}"


def analyze_document_content(doc_text: str, file_name: str, caption_prompt: str = "") -> str:
    """Sends extracted document text to Gemini 2.5 Flash for deep analysis and copyable formatting."""
    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        return "Error: GEMINI_API_KEY not configured for document processing."

    default_prompt = (
        f"You are Jarvis 3.0, the user's trusted AI partner and document intelligence analyst.\n"
        f"Analyze this uploaded document ({file_name}).\n"
        f"1. 📋 VERBATIM EXTRACTED CONTENT (Ready to copy in 1 click in a clean format).\n"
        f"2. 📊 STRUCTURED TABLE / SUMMARY OF METRICS.\n"
        f"3. 💡 KEY TAKEAWAYS & 3 ACTIONABLE RECOMMENDATIONS."
    )
    prompt = caption_prompt if caption_prompt.strip() else default_prompt

    system_instruction = (
        "You are Jarvis 3.0, an expert document analyst, OCR transcriber, accountant, and engineer. "
        "Extract text with 100% fidelity. Always provide a copyable content block and 3 strategic recommendations."
    )

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
    payload = {
        "system_instruction": {"parts": {"text": system_instruction}},
        "contents": [
            {
                "parts": [
                    {"text": f"{prompt}\n\nDOCUMENT CONTENT:\n{doc_text[:50000]}"}
                ]
            }
        ]
    }

    try:
        resp = requests.post(url, json=payload, timeout=40)
        data = resp.json()
        if "candidates" in data and len(data["candidates"]) > 0:
            return f"📄 *DOCUMENT & OCR INTELLIGENCE ({file_name})*\n\n" + data["candidates"][0]["content"]["parts"][0]["text"]
        return f"Document text extracted ({len(doc_text)} chars), but analysis failed."
    except Exception as e:
        return f"Document analysis error: {str(e)}"
